#!/usr/bin/env python3
"""
GitLab Issues Exporter
Console application to fetch GitLab issues and export them to Excel
"""

import json
import requests
from datetime import datetime
from dateutil import parser
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import sys
import os


class GitLabIssuesExporter:
    def __init__(self, config_file='config.json'):
        self.config = self.load_config(config_file)
        self.headers = {
            'Private-Token': self.config['gitlab_token'],
            'Content-Type': 'application/json'
        }
        self.base_url = self.config.get('gitlab_url', 'https://gitlab.com')
        
    def load_config(self, config_file):
        """Load configuration from JSON file"""
        try:
            with open(config_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except FileNotFoundError:
            print(f"Error: Configuration file '{config_file}' not found.")
            sys.exit(1)
        except json.JSONDecodeError:
            print(f"Error: Invalid JSON in configuration file '{config_file}'.")
            sys.exit(1)
    
    def fetch_issues(self, start_date=None, end_date=None):
        """Fetch issues from GitLab API within date range"""
        project_id = self.config['project_id']
        url = f"{self.base_url}/api/v4/projects/{project_id}/issues"
        
        params = {
            'per_page': 100,
            'page': 1,
            'state': 'all'
        }
        
        if start_date:
            params['created_after'] = start_date.isoformat()
        if end_date:
            params['created_before'] = end_date.isoformat()
        
        all_issues = []
        
        while True:
            try:
                response = requests.get(url, headers=self.headers, params=params)
                response.raise_for_status()
                
                issues = response.json()
                if not issues:
                    break
                    
                all_issues.extend(issues)
                params['page'] += 1
                
                print(f"Fetched {len(all_issues)} issues so far...")
                
            except requests.exceptions.RequestException as e:
                print(f"Error fetching issues: {e}")
                return []
        
        return all_issues
    
    def export_to_excel(self, issues, filename=None):
        """Export issues to Excel file"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"gitlab_issues_{timestamp}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "GitLab Issues"
        
        # Headers
        headers = [
            "ID del issue",
            "Título del issue", 
            "Descripción del issue",
            "Nombre del autor",
            "Estado del issue",
            "Asignados al issue",
            "Etiquetas del issue",
            "Fecha y hora de creación",
            "Tiempo total estimado",
            "Tiempo total gastado"
        ]
        
        # Add headers with formatting
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add data
        for row, issue in enumerate(issues, 2):
            # Extract assignees
            assignees = ", ".join([assignee['name'] for assignee in issue.get('assignees', [])])
            
            # Extract labels
            labels = ", ".join(issue.get('labels', []))
            
            # Format creation date
            created_at = parser.parse(issue['created_at']).strftime("%Y-%m-%d %H:%M:%S")
            
            # Time tracking
            time_stats = issue.get('time_stats', {})
            estimated_time = self.format_time_seconds(time_stats.get('time_estimate', 0))
            spent_time = self.format_time_seconds(time_stats.get('total_time_spent', 0))
            
            data = [
                issue['iid'],
                issue['title'],
                issue.get('description', ''),
                issue['author']['name'],
                issue['state'],
                assignees,
                labels,
                created_at,
                estimated_time,
                spent_time
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filename)
        return filename
    
    def format_time_seconds(self, seconds):
        """Convert seconds to human readable format"""
        if seconds == 0:
            return "0"
        
        hours = seconds // 3600
        minutes = (seconds % 3600) // 60
        remaining_seconds = seconds % 60
        
        if hours > 0:
            return f"{hours}h {minutes}m {remaining_seconds}s"
        elif minutes > 0:
            return f"{minutes}m {remaining_seconds}s"
        else:
            return f"{remaining_seconds}s"


def get_date_input(prompt):
    """Get date input from user"""
    while True:
        date_str = input(prompt).strip()
        if not date_str:
            return None
        try:
            return datetime.strptime(date_str, "%Y-%m-%d")
        except ValueError:
            print("Formato de fecha inválido. Use YYYY-MM-DD (ejemplo: 2024-01-15)")


def main():
    print("GitLab Issues Exporter")
    print("=" * 30)
    
    # Check if config file exists
    if not os.path.exists('config.json'):
        print("Error: config.json file not found. Please create it first.")
        return
    
    exporter = GitLabIssuesExporter()
    
    print("Ingrese el rango de fechas para filtrar los issues:")
    print("(Deje en blanco para no filtrar por fecha)")
    
    start_date = get_date_input("Fecha de inicio (YYYY-MM-DD): ")
    end_date = get_date_input("Fecha de fin (YYYY-MM-DD): ")
    
    print("\nObteniendo issues de GitLab...")
    issues = exporter.fetch_issues(start_date, end_date)
    
    if not issues:
        print("No se encontraron issues en el rango de fechas especificado.")
        return
    
    print(f"Se encontraron {len(issues)} issues.")
    
    # Export to Excel
    print("Exportando a Excel...")
    filename = exporter.export_to_excel(issues)
    
    print(f"Issues exportados exitosamente a: {filename}")


if __name__ == "__main__":
    main()